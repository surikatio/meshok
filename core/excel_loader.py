"""Чтение и запись списков ссылок на фотографии в Excel-файле."""

import logging
import os
import openpyxl as excel
from core.paths import get_app_dir

logger = logging.getLogger(__name__)


def save_failed_rows(table_name: str, failed_rows: list[list[str]]) -> str | None:
    """Сохраняет не выставленные строки в отдельный xlsx рядом с исходным файлом.

    Возвращает путь к созданному файлу или None при ошибке.
    """
    if not failed_rows:
        return None
    if not os.path.isabs(table_name):
        table_name = os.path.join(get_app_dir(), table_name)
    base, ext = os.path.splitext(table_name)
    out_path = base + "_failed" + (ext or ".xlsx")
    try:
        wb = excel.Workbook()
        ws = wb.active
        for row in failed_rows:
            ws.append(row)
        wb.save(out_path)
        logger.info("Failed rows saved to %s", out_path)
        return out_path
    except Exception as e:
        logger.error("Failed to save failed rows to %s: %s", out_path, e)
        return None


def load_url_list(table_name: str) -> list[list[str]]:
    """Загружает таблицу с URL картинок.

    Одна строка Excel = один лот, каждая непустая ячейка строки = URL одной фотографии.
    table_name может быть относительным (ищется рядом с приложением) или абсолютным путём.
    Возвращает пустой список, если файл не найден или его не удалось прочитать.
    """
    if not os.path.isabs(table_name):
        table_name = os.path.join(get_app_dir(), table_name)
    try:
        wb = excel.load_workbook(table_name)
        sheet = wb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            urls = [str(cell) for cell in row if cell is not None and str(cell).strip() not in ("", "None")]
            if urls:
                rows.append(urls)
        return rows
    except FileNotFoundError:
        logger.warning("Excel file not found: %s", table_name)
        return []
    except Exception as e:
        logger.error("Failed to load Excel file %s: %s", table_name, e)
        return []
